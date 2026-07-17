"""XLS/XLSX file reader — read contacts from Excel sheets.
Supports both .xls (xlrd) and .xlsx (openpyxl) formats.
"""
from modules.config import XLS_FILE, ALL_CONTACTS_FILE
from modules.validation import is_valid_email


def _open_workbook(path):
    """Open workbook — auto-detect format by extension"""
    ext = path.suffix.lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook
        # Use read_only=False for writing later (merge_xls_into_all)
        return load_workbook(str(path))
    else:
        import xlrd
        return xlrd.open_workbook(str(path))


def _get_sheet(wb):
    """Get target sheet (Namecards) or first sheet"""
    import xlrd
    from openpyxl import load_workbook

    if isinstance(wb, xlrd.Book):
        sheet_names = wb.sheet_names()
    else:
        sheet_names = wb.sheetnames

    target = None
    for s in sheet_names:
        if "namecard" in s.lower():
            target = s
            break
    if not target and sheet_names:
        target = sheet_names[0]
    return target


def _read_rows_xls(wb, sheet_name):
    """Yield (name, email, phone, job_title, company) from .xls workbook"""
    sheet = wb.sheet_by_name(sheet_name)
    for row_idx in range(2, sheet.nrows):
        name = str(sheet.cell_value(row_idx, 0)).strip()
        email = str(sheet.cell_value(row_idx, 1)).strip()
        phone = str(sheet.cell_value(row_idx, 2)).strip() if sheet.ncols > 2 else ""
        job_title = str(sheet.cell_value(row_idx, 3)).strip() if sheet.ncols > 3 else ""
        company = str(sheet.cell_value(row_idx, 4)).strip() if sheet.ncols > 4 else ""
        yield name, email, phone, job_title, company


def _read_rows_xlsx(wb, sheet_name):
    """Yield (name, email, phone, job_title, company) from .xlsx workbook"""
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip() if row[0] else ""
        email = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        phone = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        job_title = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        company = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        yield name, email, phone, job_title, company


def _read_rows(path):
    """Generic row reader — delegates to format-specific function"""
    wb = _open_workbook(path)
    sheet_name = _get_sheet(wb)
    if not sheet_name:
        wb.close() if hasattr(wb, 'close') else None
        return

    import xlrd
    if isinstance(wb, xlrd.Book):
        yield from _read_rows_xls(wb, sheet_name)
    else:
        yield from _read_rows_xlsx(wb, sheet_name)

    wb.close() if hasattr(wb, 'close') else None


def read_namecards():
    """Read contacts from Excel file (return [] if file not found)"""
    if not XLS_FILE.exists():
        print(f"[WARN] File tidak ditemukan: {XLS_FILE}")
        return []

    contacts = []
    for name, email, phone, job_title, company in _read_rows(XLS_FILE):
        if not name or name in ("", "FULL NAME", "0.0"):
            continue
        contacts.append({
            "name": name.title() if name.isupper() else name,
            "email": email,
            "company": company,
            "job_title": job_title,
        })

    print(f"[INFO] Ditemukan {len(contacts)} kontak")
    return contacts


def merge_xls_into_all(filepath):
    """Read Excel, merge with contacts_all.json (dedup by email), save.
    Return (total_baru, total_keseluruhan)
    """
    import json
    from modules.storage import load_merged_contacts, save_merged_contacts

    existing = load_merged_contacts()
    existing_emails = {c["email"].strip().lower() for c in existing if c.get("email")}

    baru = 0
    for name, email, phone, job_title, company in _read_rows(filepath):
        if not name or name in ("", "FULL NAME", "0.0"):
            continue
        email_lower = email.strip().lower()
        if email_lower and email_lower not in existing_emails:
            existing.append({
                "name": name.title() if name.isupper() else name,
                "email": email,
                "company": company,
                "job_title": job_title,
            })
            existing_emails.add(email_lower)
            baru += 1

    save_merged_contacts(existing)
    return baru, len(existing)
