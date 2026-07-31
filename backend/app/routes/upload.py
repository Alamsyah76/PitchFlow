"""File upload endpoint with format validation — supports .xls, .xlsx, .csv"""
from pathlib import Path
from fastapi import APIRouter, HTTPException, UploadFile, File
from . import logger, XLS_FILE

router = APIRouter()


def _read_sample_emails_xls(path, max_rows=5):
    """Read sample emails from .xls file using xlrd"""
    import xlrd
    from . import is_valid_email
    wb = xlrd.open_workbook(str(path))
    sheet = wb.sheet_by_index(0)
    samples = []
    for r in range(2, min(2 + max_rows, sheet.nrows)):
        email = str(sheet.cell_value(r, 1)).strip()
        if email:
            samples.append(email)
    return samples


def _read_sample_emails_xlsx(path, max_rows=5):
    """Read sample emails from .xlsx file using openpyxl"""
    from openpyxl import load_workbook
    from . import is_valid_email
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    samples = []
    for idx, row in enumerate(ws.iter_rows(min_row=3, max_row=2 + max_rows, values_only=True)):
        if row and len(row) > 1 and row[1]:
            email = str(row[1]).strip()
            if email:
                samples.append(email)
    wb.close()
    return samples


def _read_sample_emails_csv(path, max_rows=5):
    """Read sample emails from .csv file"""
    import csv
    from . import is_valid_email
    samples = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader, None)  # skip header
        next(reader, None)  # skip row 2
        for _ in range(max_rows):
            try:
                row = next(reader)
                if len(row) > 1 and row[1].strip():
                    samples.append(row[1].strip())
            except StopIteration:
                break
    return samples


@router.post("/upload")
async def upload_contacts_file(file: UploadFile = File(...)):
    try:
        ext = Path(file.filename).suffix.lower()
        if ext not in (".xls", ".xlsx", ".csv"):
            raise HTTPException(status_code=400, detail="Only .xls, .xlsx, .csv files are supported")

        content = await file.read()

        # Simpan dengan ekstensi asli — biarkan XLS_FILE tetap .xls untuk backward compat
        # Tapi simpan juga file asli dengan ekstensi aslinya
        from modules.config import REPORT_DIR
        REPORT_DIR.mkdir(parents=True, exist_ok=True)   # Pastikan folder ada
        saved_path = REPORT_DIR / f"uploaded_contacts{ext}"
        saved_path.write_bytes(content)

        # Update XLS_FILE global untuk titik referensi
        import modules.config as cfg
        cfg.XLS_FILE = saved_path

        # Validate format: read sample emails, check valid ratio
        from . import is_valid_email
        sample_emails = []
        try:
            if ext == ".xlsx":
                sample_emails = _read_sample_emails_xlsx(saved_path)
            elif ext == ".xls":
                sample_emails = _read_sample_emails_xls(saved_path)
            else:  # .csv
                sample_emails = _read_sample_emails_csv(saved_path)

            if sample_emails:
                valid_ratio = sum(1 for e in sample_emails if is_valid_email(e)) / len(sample_emails)
                if valid_ratio < 0.5:
                    raise HTTPException(status_code=400,
                        detail=f"Format file tidak sesuai. Kolom Email harus valid (hanya {int(valid_ratio*100)}% terdeteksi). Pastikan format: Nama, Email, No.Telp, Jabatan, Perusahaan")
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=400,
                detail="Format file tidak sesuai. Pastikan format kolom: Nama, Email, No.Telp, Jabatan, Perusahaan")

        from . import merge_xls_into_all
        baru, total = merge_xls_into_all(saved_path)

        return {
            "success": True,
            "message": f"File uploaded: {file.filename} (+{baru} baru, total {total} kontak)",
            "data": {"path": str(XLS_FILE), "filename": file.filename, "baru": baru, "total": total},
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
